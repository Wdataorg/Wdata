# -*- coding:utf-8 -*-


from .data_py import *
from .plot_pic import line_chart
from .read import *
from .all_error import (JsonError, FileTypeError)
from json import dump
from openpyxl import Workbook, load_workbook


JSON = "json"
XLSX = "xlsx"
CSV = "csv"

"""
Author: Wdataorg
Date: 2022.06.17
Project Name: Wdata
"""

"""
A database with multiple data sets that support drawing, 
These data sets are:
World population data set, World Carbon dioxide Concentration data set, 
World Number of Cities data set,
China number of population data set, 
China number of space vehicles data set......
"""

class WdataMain():
    def __init__(self, jsonname: str) -> None:
        self.DATA_LIST = data_list
        self.__jsonname = jsonname
        if not(self.__jsonname in self.DATA_LIST):
            raise JsonError("There is no dataset {}".format(self.__jsonname))

    def __dict__(self) -> dict:
        return read_dict(self.__jsonname)

    def __map_str(self, x:list)-> list:
        return list(map(lambda y:str(y), x))

    def Save_file(self, filename: str, type=JSON, Sheet='Data', RowOrColumn=True) -> None:
        """
        par:
            1. Sheet        Default 'Data'
            2. Row or Column (True or False)
        """
        if not(type in [JSON, CSV, XLSX]):
            raise FileTypeError('No file type {}'.format(type))

        if type == JSON:
            with open(f'{filename}.json', 'w+') as file:
                dump(read_dict(self.__jsonname), file)
            return True

        elif type == XLSX:
            # with open('f{filename.xlsx}', 'ab+') as file:
            #     pass
            #     # For Create File
            workbook = Workbook(filename + '.xlsx')
            workbook.save(filename +'.xlsx')
            workbook = load_workbook(filename +'.xlsx')

            sheet = workbook.create_sheet(Sheet)
            years_list = self.__map_str(list(read_dict(self.__jsonname).keys()))
            value_list = self.__map_str(list(read_dict(self.__jsonname).values()))
            
            for i in range(1,len(years_list)):
                if not RowOrColumn:
                    sheet.cell(row=i, column=1).value = years_list[i]
                    sheet.cell(row=i, column=2).value = value_list[i] 
                    sheet.column_dimensions['B'].width = 13.0
                else:
                    sheet.cell(row=1, column=i).value = years_list[i]
                    sheet.cell(row=2, column=i).value = value_list[i]
                    # print(chr(32))
                    # print(chr(i+64))
                    sheet.column_dimensions[chr(i+64)].width = 13.0
            workbook.close()
            workbook.save(filename+'.xlsx')
           
            
        elif type == CSV:
            with open(f'{filename}.csv', 'w+') as file:
                years_list = self.__map_str(list(read_dict(self.__jsonname).keys()))
                value_list = self.__map_str(list(read_dict(self.__jsonname).values()))
                file.write("{}\n{}".format(','.join(years_list), ','.join(value_list)))

    def __read_json_draw(self) -> dict:
        return read_dict_draw(self.__jsonname)

    def draw(self) -> None:
        result = self.__read_json_draw()
        if result[1] == 'Line Chart':
            print('Plot the data for data {}'.format(self.__jsonname))
            line_chart(result[0], self.__jsonname)
    def __len__(self) -> int:
        return len(self.__dict__())

    def __repr__(self) -> str:
        """
        eg.
        >>> from Wdata import WdataMain
        >>> test = WdataMain('Population_growth')
        >>> test
         Wdata_class('Population_growth')
        """
        return '<WdataMain jsonname="{}" Data={}>'.format(self.__jsonname, dict(self))

    def __list__(self):
        return list(self.__dict__().items())

    def __eq__(self, other):
        if self.__jsonname == other.__jsonname:
            return True

    def __hash__(self):
        return hash(self.__jsonname)

    def __getitem__(self, item: int) -> list:
        """
        eg.
        >>> from Wdata import WdataMain
        >>> test = WdataMain('Population_growth')
        >>> test[0]
        ("1800", 900000000)
        >>> for year, people in test:
        ...     print("In year:{}, people_growth:{}".format(year, people))
        In year:1800, people_growth:900000000
        In year:1820, people_growth:1100000000
        In year:1840, people_growth:1200000000
        In year:1860, people_growth:1300000000
        In year:1880, people_growth:1400000000
        In year:1900, people_growth:1650000000
        In year:1920, people_growth:1800000000
        In year:1940, people_growth:2200000000
        In year:1960, people_growth:3000000000
        In year:1980, people_growth:4400000000
        In year:2000, people_growth:5900000000
        In year:2022, people_growth:7400000000
        """
        return list(self.__dict__().items())[item]